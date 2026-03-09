"""One-shot script: read Excel knowledge base → embed → upsert to Pinecone."""
import hashlib
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from embedding.openai_embedder import embed_texts
from vectorstore.pinecone_store import upsert_vectors

EXCEL_PATH = "/Users/bytedance/Downloads/企微AI知识库（统一版）Z5.0.xlsx"


def load_qa_from_excel(path: str) -> list[dict]:
    """Read Q&A pairs from Excel. Columns: A=模块, B=类型, C=标准问题, D=参考答案."""
    wb = openpyxl.load_workbook(path)
    qa_pairs = []

    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue
            module = str(row[0]).strip() if row[0] else ""
            category = str(row[1]).strip() if row[1] else ""
            question = str(row[2]).strip() if row[2] else ""
            answer = str(row[3]).strip() if row[3] else ""
            if question and answer:
                qa_pairs.append({
                    "module": module,
                    "category": category,
                    "question": question,
                    "answer": answer,
                })

    wb.close()
    print(f"Loaded {len(qa_pairs)} Q&A pairs from {len(wb.sheetnames)} sheets")
    return qa_pairs


def main():
    # 1. Load Q&A pairs from Excel
    qa_pairs = load_qa_from_excel(EXCEL_PATH)
    if not qa_pairs:
        print("No Q&A pairs found. Exiting.")
        return

    # 2. Create embeddings (batch by 50 to avoid request size limits)
    texts = [f"Q: {p['question']}\nA: {p['answer']}" for p in qa_pairs]
    print(f"Embedding {len(texts)} texts...")

    batch_size = 50
    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        print(f"  Batch {i // batch_size + 1}/{(len(texts) - 1) // batch_size + 1} ({len(batch)} texts)")
        all_embeddings.extend(embed_texts(batch))

    # 3. Prepare vectors with MD5 IDs for idempotent upserts
    vectors = []
    for pair, embedding in zip(qa_pairs, all_embeddings):
        vec_id = hashlib.md5(pair["question"].encode()).hexdigest()
        vectors.append({
            "id": vec_id,
            "values": embedding,
            "metadata": {
                "module": pair["module"],
                "category": pair["category"],
                "question": pair["question"],
                "answer": pair["answer"],
            },
        })

    # 4. Upsert to Pinecone
    upsert_vectors(vectors)
    print("Ingestion complete!")


if __name__ == "__main__":
    main()
